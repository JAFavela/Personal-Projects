# -*- coding: utf-8 -*-
"""
Created on Fri May 22 11:15:06 2020

@name: Automatic Budget Transfer Tool
@author: Jorge Favela

"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from wx import App,FD_FILE_MUST_EXIST, FD_OPEN, FileDialog, ID_OK
import sys
import tkinter as tk
import tkinter.ttk as ttk

py3 = True
cellF = {True : 'General', False : '0.00'}
global rnd
global oName

class Toplevel1:
    global top
    def __init__(self, top=None):
        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''
        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
        _fgcolor = '#000000'  # X11 color: 'black'
        _compcolor = '#d9d9d9' # X11 color: 'gray85'
        _ana1color = '#d9d9d9' # X11 color: 'gray85'
        _ana2color = '#ececec' # Closest X11 color: 'gray92'
        self.style = ttk.Style()
        if sys.platform == "win32":
            self.style.theme_use('winnative')
        self.style.configure('.',background=_bgcolor)
        self.style.configure('.',foreground=_fgcolor)
        self.style.configure('.',font="TkDefaultFont")
        self.style.map('.',background=
            [('selected', _compcolor), ('active',_ana2color)])

        top.geometry("600x166+707+446")
        top.minsize(148, 1)
        top.maxsize(1924, 1055)
        top.resizable(0,  0)
        top.title("ABTT by Jorge Favela")
        top.configure(background="#d9d9d9")

        self.outName = ttk.Entry(top)
        self.outName.place(x=20, y=60, height=36, width=246)
        self.outName.configure(takefocus="")
        self.outName.configure(cursor="ibeam")

        self.TLabel1 = ttk.Label(top)
        self.TLabel1.place(x=20, y=30, height=24, width=185)
        self.TLabel1.configure(background="#d9d9d9")
        self.TLabel1.configure(foreground="#000000")
        self.TLabel1.configure(font="TkDefaultFont")
        self.TLabel1.configure(relief="flat")
        self.TLabel1.configure(anchor='w')
        self.TLabel1.configure(justify='left')
        self.TLabel1.configure(text='''Enter name for output file:''')

        self.TLabel2 = ttk.Label(top)
        self.TLabel2.place(x=320, y=30, height=24, width=213)
        self.TLabel2.configure(background="#d9d9d9")
        self.TLabel2.configure(foreground="#000000")
        self.TLabel2.configure(font="TkDefaultFont")
        self.TLabel2.configure(relief="flat")
        self.TLabel2.configure(anchor='w')
        self.TLabel2.configure(justify='left')
        self.TLabel2.configure(text='''Round to whole number? (Y/N):''')

        self.rund = ttk.Entry(top)
        self.rund.place(x=320, y=60, height=36, width=256)
        self.rund.configure(takefocus="")
        self.rund.configure(cursor="ibeam")

        self.TButton1 = ttk.Button(top)
        self.TButton1.place(x=20, y=110, height=30, width=558)
        self.TButton1.configure(takefocus="")
        self.TButton1.configure(text='''Continue''')
        self.TButton1.configure(command=setValues)


def getFilePath(wildcard, title):
    app = App(None)
    style = FD_OPEN | FD_FILE_MUST_EXIST
    dialog = FileDialog(None, title, wildcard=wildcard, style=style)
    if dialog.ShowModal() == ID_OK:
        path = dialog.GetPath()
    else:
        path = None
    dialog.Destroy()
    return path

def convertData(filename,kc):
    lines=[]
    dsf = load_workbook(filename=filename)
    sheet=dsf.active
    objs=sheet[kc]
    rules=[i.value for i in objs]
    cnt=0
    for line in rules:
        cnt+=1
        if line == None:
            break
        if line[0]=='-':
            lines.append([cnt,['0']])
        else:
            AN=[]
            if len(line)==11:
                AN=[line[0:11]]
                lines.append([cnt,AN])
            elif len(line)==9:
                AN=[line[0:9]]
                lines.append([cnt,AN])
            else:
                num=line.split(' ')
                for i in num:
                    if i=='\n' or i=='':
                        break
                    if i[0]=='A':
                        AN.append('Account Number')
                        break
                    w=i[7:]
                    st=i[0:7]
                    en=''
                    for j in range(len(w)):
                        if w[j]=='/':
                            AN.append(st+en)
                            en=''
                        else:
                            en=en+w[j]
                        if j==(len(w)-1):
                            AN.append(st+en)
                            en=''
                lines.append([cnt,AN])
    return lines

def compare(lines,Data,rnd):
    vals=[0.00 for i in range(len(lines)+1)]
    for rec in Data[::-1]:
        for l in lines:
            if rec[0] in l[1]:
                vals[l[0]]+=rec[1]
                break
    if rnd:
        return [round(num,0) for num in vals]
    return [round(num,2) for num in vals]

def stripSpace(s):
    s = s.replace(' ', '')
    return s.replace('\n', '')
    
        
def getNewData(datFile):
    workbook = load_workbook(filename=datFile)
    sheet = workbook.active
    Data=[]
    for value in sheet.iter_rows(max_col=2,values_only=True):
        valNoSpace=(stripSpace(value[0]),value[1])
        Data.append(valNoSpace)
    return Data

def insert2sheet(values,filename,outName,ac,rnd):
    dsf = load_workbook(filename=filename)
    sheet=dsf.active
    col=ac
    row=-0
    for amt in values:
        if row>1:
            cell=col+str(row)
            sheet[cell].value=float(amt)
            sheet[cell].alignment = Alignment(horizontal="right")
            sheet[cell].number_format = cellF.get(rnd)
        row+=1
    dsf.save(outName)

def setValues():
    global oName
    global rnd
    global top
    global root
    oName=top.outName.get() + '.xlsx'
    rnd= top.rund.get()
    if rnd.lower() == 'y':
        rnd= True
    else:
        rnd=False
    root.quit()
    
 
if __name__ == "__main__":
    global oName
    global rnd
    global top
    global root
    
    choice= lambda r: 'Yes' if r else 'No'
    print('Select the DFA template that includes the keys') # QR.xlsx
    filename = getFilePath('*.xlsx','Select the DFA template that includes the keys' )
    print('Selected File: ', filename)

    print('Select the file that contains the dollar amounts to be processed') # QR.xlsx
    datFile = getFilePath('*.xlsx', 'Select the file that contains the dollar amounts to be processed: ')
    print('Selected File: ', datFile)
    
    root = tk.Tk()
    top = Toplevel1(root)
    root.mainloop()
    on = oName
    r=rnd
    root.destroy()
    root=None

    print('Enter name to give the newly generated file')
    print('Output File Name: ', on)
    
    print('ERound final values to whole number? Y/N')
    print('Round: ', choice(r))
    
    values=compare(convertData(filename,'D'),getNewData(datFile),r)
    insert2sheet(values,filename,on,'D',r)
    
    print("\nSuccess! Your file has been saved as {} \nHope it helped!".format(on))
    